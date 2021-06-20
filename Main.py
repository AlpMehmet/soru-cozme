tc = " "
sifre=" "
secili=" "
dogruSay=0
yanlisSay=0
gelenSoruNo=0
kullaniciListe= ["","","","",""]
dogruCevap=" "
toggleDeger=0   
soruSay=0
seciliDers= ""
import time
import random
import sys
import threading 
from  openpyxl import *
from PyQt5 import QtCore, QtGui, QtWidgets
from Bilgiler import Ui_Bilgiler
class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(758, 850)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(MainWindow.sizePolicy().hasHeightForWidth())
        MainWindow.setSizePolicy(sizePolicy)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.verticalLayoutWidget = QtWidgets.QWidget(self.centralwidget)
        self.verticalLayoutWidget.setGeometry(QtCore.QRect(210, 20, 281, 151))
        self.verticalLayoutWidget.setObjectName("verticalLayoutWidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.verticalLayoutWidget)
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout.setObjectName("verticalLayout")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.labelSifre = QtWidgets.QLabel(self.verticalLayoutWidget)
        self.labelSifre.setObjectName("labelSifre")
        self.horizontalLayout_2.addWidget(self.labelSifre)
        self.lineEditSifre = QtWidgets.QLineEdit(self.verticalLayoutWidget)
        self.lineEditSifre.setObjectName("lineEditSifre")
        self.horizontalLayout_2.addWidget(self.lineEditSifre)
        self.verticalLayout.addLayout(self.horizontalLayout_2)
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.labelTc = QtWidgets.QLabel(self.verticalLayoutWidget)
        self.labelTc.setObjectName("labelTc")
        self.horizontalLayout.addWidget(self.labelTc)
        self.lineEditTc = QtWidgets.QLineEdit(self.verticalLayoutWidget)
        self.lineEditTc.setObjectName("lineEditTc")
        self.horizontalLayout.addWidget(self.lineEditTc)
        self.verticalLayout.addLayout(self.horizontalLayout)
        self.pushButtonGiris = QtWidgets.QPushButton(self.verticalLayoutWidget)
        self.pushButtonGiris.setObjectName("pushButtonGiris")
        self.verticalLayout.addWidget(self.pushButtonGiris)
        self.groupBox = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox.setGeometry(QtCore.QRect(20, 190, 671, 561))
        self.groupBox.setObjectName("groupBox")
        self.verticalLayoutWidget_2 = QtWidgets.QWidget(self.groupBox)
        self.verticalLayoutWidget_2.setGeometry(QtCore.QRect(10, 60, 661, 441))
        self.verticalLayoutWidget_2.setObjectName("verticalLayoutWidget_2")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_2)
        self.verticalLayout_2.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.horizontalLayout_7 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_7.setObjectName("horizontalLayout_7")
        self.labelSoru = QtWidgets.QLabel(self.verticalLayoutWidget_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.labelSoru.sizePolicy().hasHeightForWidth())
        self.labelSoru.setSizePolicy(sizePolicy)
        self.labelSoru.setObjectName("labelSoru")
        self.horizontalLayout_7.addWidget(self.labelSoru)
        self.verticalLayout_2.addLayout(self.horizontalLayout_7)
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.label_7 = QtWidgets.QLabel(self.verticalLayoutWidget_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_7.sizePolicy().hasHeightForWidth())
        self.label_7.setSizePolicy(sizePolicy)
        self.label_7.setObjectName("label_7")
        self.horizontalLayout_3.addWidget(self.label_7)
        self.labelCevap1 = QtWidgets.QLabel(self.verticalLayoutWidget_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.labelCevap1.sizePolicy().hasHeightForWidth())
        self.labelCevap1.setSizePolicy(sizePolicy)
        self.labelCevap1.setObjectName("labelCevap1")
        self.horizontalLayout_3.addWidget(self.labelCevap1)
        self.verticalLayout_2.addLayout(self.horizontalLayout_3)
        self.horizontalLayout_5 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_5.setObjectName("horizontalLayout_5")
        self.label_10 = QtWidgets.QLabel(self.verticalLayoutWidget_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_10.sizePolicy().hasHeightForWidth())
        self.label_10.setSizePolicy(sizePolicy)
        self.label_10.setObjectName("label_10")
        self.horizontalLayout_5.addWidget(self.label_10)
        self.labelCevap2 = QtWidgets.QLabel(self.verticalLayoutWidget_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.labelCevap2.sizePolicy().hasHeightForWidth())
        self.labelCevap2.setSizePolicy(sizePolicy)
        self.labelCevap2.setObjectName("labelCevap2")
        self.horizontalLayout_5.addWidget(self.labelCevap2)
        self.verticalLayout_2.addLayout(self.horizontalLayout_5)
        self.horizontalLayout_4 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")
        self.label_8 = QtWidgets.QLabel(self.verticalLayoutWidget_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_8.sizePolicy().hasHeightForWidth())
        self.label_8.setSizePolicy(sizePolicy)
        self.label_8.setObjectName("label_8")
        self.horizontalLayout_4.addWidget(self.label_8)
        self.labelCevap3 = QtWidgets.QLabel(self.verticalLayoutWidget_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.labelCevap3.sizePolicy().hasHeightForWidth())
        self.labelCevap3.setSizePolicy(sizePolicy)
        self.labelCevap3.setObjectName("labelCevap3")
        self.horizontalLayout_4.addWidget(self.labelCevap3)
        self.verticalLayout_2.addLayout(self.horizontalLayout_4)
        self.horizontalLayout_6 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_6.setObjectName("horizontalLayout_6")
        self.label_12 = QtWidgets.QLabel(self.verticalLayoutWidget_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_12.sizePolicy().hasHeightForWidth())
        self.label_12.setSizePolicy(sizePolicy)
        self.label_12.setObjectName("label_12")
        self.horizontalLayout_6.addWidget(self.label_12)
        self.labelCevap4 = QtWidgets.QLabel(self.verticalLayoutWidget_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.labelCevap4.sizePolicy().hasHeightForWidth())
        self.labelCevap4.setSizePolicy(sizePolicy)
        self.labelCevap4.setObjectName("labelCevap4")
        self.horizontalLayout_6.addWidget(self.labelCevap4)
        self.verticalLayout_2.addLayout(self.horizontalLayout_6)
        self.horizontalLayout_8 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_8.setObjectName("horizontalLayout_8")
        self.radioButtonA = QtWidgets.QRadioButton(self.verticalLayoutWidget_2)
        self.radioButtonA.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.radioButtonA.setLayoutDirection(QtCore.Qt.RightToLeft)
        self.radioButtonA.setObjectName("radioButtonA")
        self.horizontalLayout_8.addWidget(self.radioButtonA)
        self.radioButtonB = QtWidgets.QRadioButton(self.verticalLayoutWidget_2)
        self.radioButtonB.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.radioButtonB.setLayoutDirection(QtCore.Qt.RightToLeft)
        self.radioButtonB.setObjectName("radioButtonB")
        self.horizontalLayout_8.addWidget(self.radioButtonB)
        self.radioButtonC = QtWidgets.QRadioButton(self.verticalLayoutWidget_2)
        self.radioButtonC.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.radioButtonC.setLayoutDirection(QtCore.Qt.RightToLeft)
        self.radioButtonC.setObjectName("radioButtonC")
        self.horizontalLayout_8.addWidget(self.radioButtonC)
        self.radioButtonD = QtWidgets.QRadioButton(self.verticalLayoutWidget_2)
        self.radioButtonD.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.radioButtonD.setLayoutDirection(QtCore.Qt.RightToLeft)
        self.radioButtonD.setObjectName("radioButtonD")
        self.horizontalLayout_8.addWidget(self.radioButtonD)
        self.verticalLayout_2.addLayout(self.horizontalLayout_8)
        self.labelSeciliDers = QtWidgets.QLabel(self.groupBox)
        self.labelSeciliDers.setGeometry(QtCore.QRect(10, 40, 91, 16))
        self.labelSeciliDers.setObjectName("labelSeciliDers")
        self.comboBoxDersler = QtWidgets.QComboBox(self.groupBox)
        self.comboBoxDersler.setGeometry(QtCore.QRect(10, 20, 661, 22))
        self.comboBoxDersler.setObjectName("comboBoxDersler")
        self.horizontalLayoutWidget_9 = QtWidgets.QWidget(self.groupBox)
        self.horizontalLayoutWidget_9.setGeometry(QtCore.QRect(10, 510, 661, 41))
        self.horizontalLayoutWidget_9.setObjectName("horizontalLayoutWidget_9")
        self.horizontalLayout_9 = QtWidgets.QHBoxLayout(self.horizontalLayoutWidget_9)
        self.horizontalLayout_9.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout_9.setObjectName("horizontalLayout_9")
        self.labelKontrol = QtWidgets.QLabel(self.horizontalLayoutWidget_9)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.labelKontrol.sizePolicy().hasHeightForWidth())
        self.labelKontrol.setSizePolicy(sizePolicy)
        self.labelKontrol.setObjectName("labelKontrol")
        self.horizontalLayout_9.addWidget(self.labelKontrol)
        self.labelZaman = QtWidgets.QLabel(self.horizontalLayoutWidget_9)
        self.labelZaman.setObjectName("labelZaman")
        self.horizontalLayout_9.addWidget(self.labelZaman)
        self.pushButtonBitir = QtWidgets.QPushButton(self.horizontalLayoutWidget_9)
        self.pushButtonBitir.setObjectName("pushButtonBitir")
        self.horizontalLayout_9.addWidget(self.pushButtonBitir)
        self.pushButtonGec = QtWidgets.QPushButton(self.horizontalLayoutWidget_9)
        self.pushButtonGec.setObjectName("pushButtonGec")
        self.horizontalLayout_9.addWidget(self.pushButtonGec)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 758, 23))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        
  
  
#        timer = threading.Timer(10100.0, self.gfg) 
#        timer.start() 
#        print("Exit\n") 
        
        self.comboBoxDersler.addItem("Matematik")
        self.comboBoxDersler.addItem("Fizik")
        self.comboBoxDersler.addItem("Bilgisayar")
        self.comboBoxDersler.addItem("Geometri")
        
        self.comboBoxDersler.activated[str].connect(self.onChanged)
        self.pushButtonGec.clicked.connect(self.yeniSoru)
        
        login = load_workbook("Login.xlsx")
        sheet=login.active
        say=0
        for row in sheet.iter_rows(min_row=2, min_col=1, max_row=2, max_col=4):
            for cell in row:
                kullaniciListe[say]=cell.value
                say=say+1
        self.tc=kullaniciListe[0]
        self.sifre=kullaniciListe[3]
        login.close()
        self.pushButtonBitir.clicked.connect(self.openWindow)
        self.pushButtonGiris.clicked.connect(self.giris)
     
        
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
#    say213=0
#    def gfg(): 
#        self.say213=self.say213+1
#        print(self.say213)     
    def onChanged(self, text):
        self.gelenSoruNo = random.randint(0,20)
        self.labelSeciliDers.setText(text)
        self.seciliDers = text
        
        soruBank = load_workbook("SoruBankası.xlsx")
        sheet = soruBank.active
        say=0
        for row in sheet.iter_rows(min_row=self.gelenSoruNo, min_col=1, max_row=self.gelenSoruNo, max_col=6):
            for cell in row:
                if(say==0):
                    self.labelSoru.setText("Soru "+str(self.gelenSoruNo)+": "+cell.value)
                elif(say==1):
                    self.labelCevap1.setText(cell.value)
                elif(say==2):
                    self.labelCevap2.setText(cell.value)
                elif(say==3):
                    self.labelCevap3.setText(cell.value)
                elif(say==4):
                    self.labelCevap4.setText(cell.value)
                elif(say==5):
                    self.dogruCevap=cell.value
                say=say+1
        self.toggleDeger=self.gelenSoruNo
        soruBank.close()
        self.comboBoxDersler.setVisible(False)
    def giris(self):

        sayK=0
        sayH=0
        sifreGir=self.lineEditSifre.text()
        tcGir=self.lineEditTc.text()

        for i in sifreGir:
            if (chr(ord(i))>=chr(65) and chr(ord(i))<= chr(90)) or (chr(ord(i))>=chr(97) and chr(ord(i))<= chr(122)):
                sayH=sayH+1
            else:
                sayK=sayK+1
        if(sayH==4 and sayK==4):
            gecici=self.sifre
            if(gecici==sifreGir):
                print(self.sifre)
                gecici2=self.tc
                print(self.tc)
                print(sifreGir)
                print(tcGir)               
                if(str(gecici2)==str(tcGir)):
                    self.groupBox.setVisible(True)
                else:
                    print("tc yanlış lütfen 11 karakterli tc kimlik numaranızı giniriz")
            else:
                print("şifre yanlış")
        else:
            print("Lütfen 4 harf ve 4 karakter giriniz")
         
    def yeniSoru(self):
        self.labelKontrol.setText("                                      ")
        self.radioButtonA.setIcon(QtGui.QIcon(""))   
        self.radioButtonB.setIcon(QtGui.QIcon("")) 
        self.radioButtonC.setIcon(QtGui.QIcon("")) 
        self.radioButtonD.setIcon(QtGui.QIcon(""))   
        self.radioButtonA.setChecked(False)
        self.radioButtonB.setChecked(False)
        self.radioButtonC.setChecked(False)
        self.radioButtonD.setChecked(False)
        self.toggleDeger=self.gelenSoruNo 
        soruBank = load_workbook("SoruBankası.xlsx")
        sheet = soruBank.active
        say=0
        for row in sheet.iter_rows(min_row=self.gelenSoruNo, min_col=1, max_row=self.gelenSoruNo, max_col=6):
            
            for cell in row:
                if(say==0):
                    self.labelSoru.setText("Soru "+str(self.gelenSoruNo)+": "+cell.value)
                elif(say==1):
                    self.labelCevap1.setText(cell.value)
                elif(say==2):
                    self.labelCevap2.setText(cell.value)
                elif(say==3):
                    self.labelCevap3.setText(cell.value)
                elif(say==4):
                    self.labelCevap4.setText(cell.value)
                elif(say==5):
                    self.dogruCevap=cell.value
                say=say+1

        soruBank.close()         

    def toggleRadio(self,b):
        
        if b.text() == "A":
            if b.isChecked() == True:
                self.secili=b.text()
            else:
                print (b.text()+" is deselected")
				
        if b.text() == "B":
            if b.isChecked() == True:
                self.secili=b.text()
            else:
                print (b.text()+" is deselected")
        if b.text() == "C":
            if b.isChecked() == True:
                self.secili=b.text()
            else:
                print (b.text()+" is deselected")
        if b.text() == "D":
            if b.isChecked() == True:
                self.secili=b.text()
            else:
                print (b.text()+" is deselected") 
        if(b.text()==self.dogruCevap):
            if(self.toggleDeger==self.gelenSoruNo):
                 while self.toggleDeger==self.gelenSoruNo:
                     self.gelenSoruNo = random.randint(0,20)
                 self.labelKontrol.setText("Cevap Doğru          ")
                 self.dogruSay=dogruSay+1
                 if b.text() == "A":
                     self.radioButtonA.setIcon(QtGui.QIcon("check.jpg"))
                 if b.text() == "B":
                     self.radioButtonB.setIcon(QtGui.QIcon("check.jpg"))
                 if b.text() == "C":
                     self.radioButtonC.setIcon(QtGui.QIcon("check.jpg"))
                 if b.text() == "D":
                     self.radioButtonD.setIcon(QtGui.QIcon("check.jpg"))
        else:
            if(self.toggleDeger==self.gelenSoruNo):
                 while self.toggleDeger==self.gelenSoruNo:
                     self.gelenSoruNo = random.randint(0,20)
                 self.labelKontrol.setText("Cevap Yanlış          ")
                 self.yanlisSay=yanlisSay+1
                 if b.text() == "A":
                     self.radioButtonA.setIcon(QtGui.QIcon("carpi.jpg"))
                 if b.text() == "B":
                     self.radioButtonB.setIcon(QtGui.QIcon("carpi.jpg"))
                 if b.text() == "C":
                     self.radioButtonC.setIcon(QtGui.QIcon("carpi.jpg"))
                 if b.text() == "D":
                     self.radioButtonD.setIcon(QtGui.QIcon("carpi.jpg")) 
    def openWindow(self):
        Bilgi = Workbook()
        sheet = Bilgi.active
        sheet.append([self.dogruSay,self.yanlisSay,self.seciliDers])
        Bilgi.save("Bilgi.xlsx")
        Bilgi.close()
        self.window = QtWidgets.QMainWindow()
        self.ui = Ui_Bilgiler()
        self.ui.setupUi(self.window)
        MainWindow.hide()
        self.window.show()      
    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.lineEditSifre.setEchoMode(QtWidgets.QLineEdit.Password)
        self.labelSifre.setText(_translate("MainWindow", "Şifre:"))
        self.labelTc.setText(_translate("MainWindow", "Tc:  "))
        self.pushButtonGiris.setText(_translate("MainWindow", "Giriş"))
        self.groupBox.setTitle(_translate("MainWindow", "Sınav Ekranı"))
        self.groupBox.setVisible(False)
        self.labelSoru.setText(_translate("MainWindow", "Soru"))
        self.label_7.setText(_translate("MainWindow", "A-"))
        self.labelCevap1.setText(_translate("MainWindow", "Cevap1"))
        self.label_10.setText(_translate("MainWindow", "B-"))
        self.labelCevap2.setText(_translate("MainWindow", "Cevap1"))
        self.label_8.setText(_translate("MainWindow", "C-"))
        self.labelCevap3.setText(_translate("MainWindow", "Cevap1"))
        self.label_12.setText(_translate("MainWindow", "D-"))
        self.labelCevap4.setText(_translate("MainWindow", "Cevap1"))
        self.radioButtonA.setText(_translate("MainWindow", "A"))
        self.radioButtonA.toggled.connect(lambda:self.toggleRadio(self.radioButtonA))
        self.radioButtonB.setText(_translate("MainWindow", "B"))
        self.radioButtonB.toggled.connect(lambda:self.toggleRadio(self.radioButtonB))
        self.radioButtonC.setText(_translate("MainWindow", "C"))
        self.radioButtonC.toggled.connect(lambda:self.toggleRadio(self.radioButtonC))
        self.radioButtonD.setText(_translate("MainWindow", "D"))
        self.radioButtonD.toggled.connect(lambda:self.toggleRadio(self.radioButtonD))
        self.labelSeciliDers.setText(_translate("MainWindow", "Seçilen Ders"))
        self.labelKontrol.setText(_translate("MainWindow", "Cevap Doğru          "))
        self.labelZaman.setText(_translate("MainWindow", "Kalan Zaman"))
        self.pushButtonBitir.setText(_translate("MainWindow", "Bitir"))
        self.pushButtonGec.setText(_translate("MainWindow", "Sonraki Soru"))
        self.radioButtonA.setAutoExclusive(False)
        self.radioButtonB.setAutoExclusive(False)
        self.radioButtonC.setAutoExclusive(False)
        self.radioButtonD.setAutoExclusive(False)

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())

